import math

from django.http import HttpResponse
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportView(View):
    filename = 'sample.xlsx'
    page_title = 'Sample'
    filters = []
    headers = []
    content = []
    default_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def filter_queryset(self, queryset):
        for filter_name in self.filters:
            filter_value = self.request.GET.get(filter_name)
            queryset = queryset.filter(**{filter_name: filter_value}) if filter_value else queryset
        return queryset

    def get_filename(self):
        return self.filename

    def get_header(self):
        return self.headers

    def get_content(self):
        return self.content

    def set_headers(self, ws):
        headers = self.get_header()
        font = Font(name='Arial', size=10, color="ffffff", bold=True)
        alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
        fill = PatternFill("solid", fgColor='E55E47')
        for i, header in enumerate(headers):
            cell = ws.cell(row=1, column=i + 1)
            cell.value = header
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = self.default_border

    def set_content(self, ws):
        content = self.get_content()
        for i, row in enumerate(content):
            for j, data in enumerate(row):
                cell = ws.cell(row=2 + i, column=j + 1)
                cell.value = data
                cell.border = self.default_border

    def fix_col_width(self, ws, ratio=1.2):
        for column_cells in ws.iter_cols():
            if not column_cells:
                continue
            length = max(len(str(cell.value or "")) for cell in column_cells)
            col = get_column_letter(column_cells[0].column)
            size = length * ratio
            if size > ws.column_dimensions[col].width:
                ws.column_dimensions[col].width = math.ceil(size)

    def create_response(self):
        filename = self.get_filename()
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = self.page_title
        self.set_headers(ws)
        self.set_content(ws)
        self.fix_col_width(ws)
        response = self.create_response()
        wb.save(response)
        return response
